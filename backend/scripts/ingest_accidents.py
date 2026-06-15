"""Ingesta del dataset oficial de accidentalidad de Medellín a PostGIS.

Fuente: Secretaría de Movilidad de Medellín — dataset abierto Mendeley r6g5dfnpgh
(CC BY 4.0), 2008-2025. Lee el XLSX, preserva los campos categóricos y persiste en
accident_incidents. Idempotente por `llave` (ON CONFLICT DO NOTHING).

Uso:
    python -m scripts.ingest_accidents /ruta/Fatal_Road_Traffic.xlsx
"""
import sys
from datetime import date, datetime

import openpyxl
from sqlalchemy import text

from app.db.database import async_session_maker
import asyncio

BATCH = 2000


def _parse_date(v):
    if isinstance(v, (datetime, date)):
        return v if isinstance(v, date) and not isinstance(v, datetime) else v.date()
    return None


async def _flush(db, batch):
    await db.execute(
        text(
            "INSERT INTO accident_incidents "
            "(llave, year, incident_date, incident_hour, incident_class, severity, comuna, barrio, geom) "
            "VALUES (:llave,:year,:incident_date,:incident_hour,:incident_class,:severity,:comuna,:barrio,"
            "CASE WHEN CAST(:lng AS float8) IS NULL OR CAST(:lat AS float8) IS NULL THEN NULL "
            "ELSE ST_SetSRID(ST_MakePoint(CAST(:lng AS float8), CAST(:lat AS float8)),4326) END) "
            "ON CONFLICT (llave) DO NOTHING"
        ),
        batch,
    )
    await db.commit()


async def ingest(path: str) -> int:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    total_rows = ws.max_row - 1  # -1 por el header
    rows = ws.iter_rows(values_only=True)
    headers = next(rows)
    idx = {h: i for i, h in enumerate(headers)}

    def col(r, name):
        i = idx.get(name)
        return r[i] if i is not None and i < len(r) else None

    def s(v):
        if v is None or v == "":
            return None
        return str(v)

    inserted = 0
    batch = []
    last_log = 0
    log_interval = max(total_rows // 20, 1)
    async with async_session_maker() as db:
        for i, r in enumerate(rows, 1):
            llave = col(r, "LLAVE")
            if llave is None:
                continue
            lat, lng = col(r, "Latitud"), col(r, "Longitud")
            batch.append({
                "llave": str(llave),
                "year": col(r, "AÑO"),
                "incident_date": _parse_date(col(r, "FECHA_INCIDENTE")),
                "incident_hour": str(col(r, "HORA_INCIDENTE") or "")[:20] or None,
                "incident_class": s(col(r, "CLASE_INCIDENTE")),
                "severity": s(col(r, "GRAVEDAD_INCIDENTE")),
                "comuna": s(col(r, "Comuna Planeacion")),
                "barrio": s(col(r, "Barrio Planeacion")),
                "lat": float(lat) if lat not in (None, "", "N/D") else None,
                "lng": float(lng) if lng not in (None, "", "N/D") else None,
            })
            if i - last_log >= log_interval:
                pct = i * 100 // total_rows
                print(f"   Progreso: {pct}% ({i:,}/{total_rows:,})")
                last_log = i
            if len(batch) >= BATCH:
                await _flush(db, batch)
                inserted += len(batch)
                batch = []
        if batch:
            await _flush(db, batch)
            inserted += len(batch)
    wb.close()
    return inserted


if __name__ == "__main__":
    p = sys.argv[1] if len(sys.argv) > 1 else "Fatal_Road_Traffic.xlsx"
    n = asyncio.run(ingest(p))
    print(f"Ingesta OK: {n} incidentes procesados desde {p}")
